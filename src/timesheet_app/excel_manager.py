"""Вспомогательные функции для работы с Excel.

Содержит:
- константы имён листов;
- загрузку справочников (проекты и виды работ);
- добавление записи о затраченном времени;
- создание шаблонной книги с нужными листами и заголовками.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import Workbook, load_workbook


# Имена листов в книге Excel
REFERENCE_SHEET = "Справочник"
TIMESHEET_SHEET = "Учет времени"


class ExcelStructureError(RuntimeError):
    """Структура книги Excel не соответствует ожиданиям."""


def _normalise(values: Iterable[str | None]) -> List[str]:
    """Очистка и нормализация значений (удаляем пустые, дубликаты)."""

    items: List[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text not in items:
            items.append(text)
    return items


def load_reference_data(path: Path | str) -> Tuple[List[str], List[str]]:
    """Прочитать лист справочника и вернуть два списка: проекты и виды работ."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    if REFERENCE_SHEET not in workbook:
        raise ExcelStructureError(
            f"Workbook must contain sheet '{REFERENCE_SHEET}'. Found: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[REFERENCE_SHEET]

    projects: List[str] = []
    work_types: List[str] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Пропускаем возможную строку заголовков
        if idx == 1:
            continue
        project, work_type, *_ = row + (None, None)
        projects.append(project)
        work_types.append(work_type)

    return _normalise(projects), _normalise(work_types)


def append_time_entry(
    path: Path | str,
    *,
    project: str,
    work_type: str,
    elapsed_seconds: float,
    finished_at: datetime | None = None,
) -> None:
    """Добавить строку на лист учёта времени (дата, проект, вид работ, длительность).

    В отличие от простого `sheet.append`, мы ищем первую по-настоящему пустую
    строку, игнорируя форматирование (цвета, границы) и удалённые строки.
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path)

    if TIMESHEET_SHEET not in workbook:
        raise ExcelStructureError(
            f"Workbook must contain sheet '{TIMESHEET_SHEET}'. Found: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[TIMESHEET_SHEET]
    timestamp = finished_at or datetime.now()

    duration = timedelta(seconds=elapsed_seconds)

    # Поиск первой полностью пустой строки, начиная со 2-й (после заголовков)
    def is_empty_row(r: int) -> bool:
        return (
            sheet.cell(row=r, column=1).value is None
            and sheet.cell(row=r, column=2).value is None
            and sheet.cell(row=r, column=3).value is None
            and sheet.cell(row=r, column=4).value is None
        )

    first = 2
    last = sheet.max_row
    target_row = None
    for r in range(first, last + 1):
        if is_empty_row(r):
            target_row = r
            break
    if target_row is None:
        target_row = last + 1

    # Записываем значения по ячейкам — так мы не зависим от sheet.append
    sheet.cell(row=target_row, column=1).value = timestamp.date()
    sheet.cell(row=target_row, column=2).value = project
    sheet.cell(row=target_row, column=3).value = work_type
    cell_duration = sheet.cell(row=target_row, column=4)
    cell_duration.value = duration
    # Красивый формат времени часов:минуты:секунды
    try:
        cell_duration.number_format = "[h]:mm:ss"
    except Exception:
        pass

    workbook.save(workbook_path)


def create_template(path: Path | str) -> None:
    """Создать пустую книгу Excel с нужными листами и заголовками."""

    workbook_path = Path(path)
    wb = Workbook()
    # Удалим дефолтный лист, чтобы контролировать порядок
    default = wb.active
    wb.remove(default)

    # Лист справочника
    ws_ref = wb.create_sheet(REFERENCE_SHEET)
    ws_ref.append(["Проект", "Вид работ"])  # заголовки

    # Лист учёта времени
    ws_ts = wb.create_sheet(TIMESHEET_SHEET)
    ws_ts.append(["Дата", "Проект", "Вид работ", "Длительность"])  # заголовки

    wb.save(workbook_path)
